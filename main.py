#!/usr/bin/env python3

# Libraries 
import time
from openpyxl import Workbook
import sys


# For website parser
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

options = Options()
options.add_argument("start-maximized")
options.add_argument("disable-infobars")
options.add_argument("--disable-extensions")
options.add_argument("--disable-dev-shm-usage")
options.add_argument("--no-sandbox")
options.add_argument('--headless')
#options.add_argument('--disable-gpu')
options.add_argument("--user-data-dir=C:\\Users\\knham\\AppData\\Local\\Google\\Chrome\\User Data") #e.g. C:\Users\You\AppData\Local\Google\Chrome\User Data
options.add_argument("profile-directory=Profile 1")
prefs = {"profile.managed_default_content_settings.images": 2}
options.add_experimental_option("prefs", prefs)
browser = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)


# Standard path for employee search 
vman_employee_path = "https://www.virtualmanager.com/employees/search?speciality=coach&country_id=&job_status=1&age_min=&age_max=45&search=1&commit=S%C3%B8g"


# Only use this if you are not using chrome profiles
def vman_login():
    # Login on the webpage
    id_box = browser.find_element(By.ID, 'email')
    id_box.send_keys('')
    pw_box = browser.find_element(By.NAME, 'password')
    pw_box.send_keys('')
    # Press login button
    browser.find_element(By.NAME, "loginbtn").click()
    time.sleep(1)

    # Get cookies
    print(browser.get_cookies())



employee_list = {
    "name": [],
    "age": [],
    "value": [],
    "link": [],
    "stats": {
        "Ungdomsspillere": [],
        "Målmandstræning": [],
        "Markspillertræning": [],
        "Disciplin": [],
        "Potentialebedømmelse": [],
        "Ledelse": [],
        "Egenskabsbedømmelse": [],
        "Motivation": []
    }

}


if __name__=='__main__':
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Vman Træner"
    #browser.get(vman_employee_path)
    #vman_login()

    try:
        # Go though each page
        for k in range (893):
            print(f"On page {1+k}")
            #browser.get(f"https://www.virtualmanager.com/employees/search?age_max=36&age_min=&commit=S%C3%B8g&country_id=&job_status=1&page={k+1}&search=1&speciality=coach")
            #browser.get(f"https://www.virtualmanager.com/employees/search?age_max=45&age_min=40&commit=S%C3%B8g&country_id=&job_status=1&page={k+1}&search=1&speciality=coach")
            #browser.get(f"https://www.virtualmanager.com/employees/search?age_max=51&age_min=46&commit=S%C3%B8g&country_id=&job_status=1&page={k+1}&search=1&speciality=coach")
            browser.get(f"https://www.virtualmanager.com/employees/search?age_max=57&age_min=52&commit=S%C3%B8g&country_id=&job_status=1&page={k+1}&search=1&speciality=coach")
            # List all employee on page
            for i in range(2,27):
                employee = browser.find_element(By.CSS_SELECTOR, f"table.stretch:nth-child(2) > tbody:nth-child(1) > tr:nth-child({i})")
                words = employee.text.split()
                employee_list["name"] = words[0] + " " + words[1]
                employee_list["age"] = words[3]
                employee_list["wage"] = words[6]
                employee_list["link"] = browser.find_element(By.PARTIAL_LINK_TEXT, words[0]).get_attribute("href")


                # Find all stats
                for j in range(1,9):
                    stat_name = browser.find_element(By.CSS_SELECTOR, f"table.stretch:nth-child(2) > tbody:nth-child(1) > tr:nth-child(2) > td:nth-child(8) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child({j}) > td:nth-child(1)").get_attribute("innerHTML")
                    employee_list["stats"][stat_name] = browser.find_element(By.CSS_SELECTOR, f"table.stretch:nth-child(2) > tbody:nth-child(1) > tr:nth-child({i}) > td:nth-child(8) > div:nth-child(1) > div:nth-child(1) > div:nth-child(1) > div:nth-child(2) > table:nth-child(1) > tbody:nth-child(1) > tr:nth-child({j}) > td:nth-child(2)").get_attribute("innerHTML")
            
            
            
                # Save player to excel
                sheet.cell(row=1+i+(k*25), column=1).hyperlink = employee_list["link"]
                sheet.cell(row=1+i+(k*25), column=1).value = employee_list["name"]
                sheet.cell(row=1+i+(k*25), column=1).style = "Hyperlink"
                sheet.cell(row=1+i+(k*25), column=2).value = employee_list["age"]
                sheet.cell(row=1+i+(k*25), column=3).value = employee_list["wage"]
                sheet.cell(row=1+i+(k*25), column=4).value = employee_list["stats"]["Ungdomsspillere"]
                sheet.cell(row=1+i+(k*25), column=5).value = employee_list["stats"]["Målmandstræning"]
                sheet.cell(row=1+i+(k*25), column=6).value = employee_list["stats"]["Markspillertræning"]
                sheet.cell(row=1+i+(k*25), column=7).value = employee_list["stats"]["Disciplin"]
                sheet.cell(row=1+i+(k*25), column=8).value = employee_list["stats"]["Potentialebedømmelse"]
                sheet.cell(row=1+i+(k*25), column=9).value = employee_list["stats"]["Ledelse"]
                sheet.cell(row=1+i+(k*25), column=10).value = employee_list["stats"]["Egenskabsbedømmelse"]
                sheet.cell(row=1+i+(k*25), column=11).value = employee_list["stats"]["Motivation"]
                
            # Save the excel sheet every 100 pages
            if k % 100 == 0:
                wb.save("vman_employee.xlsx")
                print("Saving data")
        browser.quit()
        sys.exit()
    except:
        wb.save("vman_employee.xlsx") 
        browser.quit()
        sys.exit()
    








